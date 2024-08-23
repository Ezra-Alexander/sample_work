import numpy as np
import sys
import math
from geom_helper import *
from openpyxl import Workbook, load_workbook
from qchem_helper import get_geom_io

def main():

	# for a set of serial single points, extracts specified MO energies and writes them to an excel file.
	# also extracts and writes the chelpg charges for each "center" atom

	out_file=sys.argv[1] #name of the serialized qchem.out
	excel=sys.argv[2] #name of excel file to write
	nval=int(sys.argv[3]) #number of orbitals to plot from band gap
	occ=sys.argv[4] #either o for occupied or u for unoccupied. can enter guess and it will guess

	wb=Workbook()
	ws = wb.active

	#identify "centers"
	centers=[]
	atoms,coords=get_geom_io(out_file)
	for i,atom in enumerate(atoms):
		if occ.lower()=="o":
			if atom=="P":
				centers.append(i)
		elif occ.lower()=="u":
			if atom=="In" or atom=="Ga":
				centers.append(i)
		else:
			raise Exception("Acceptable values for occ are 'o', 'u', and 'guess'")

	#make column headers
	alphabet="ABCDEFGHIJKLMNOPQRSTUVWXYZ" #lol
	ws['A1']="Frame"
	ws["B1"]="Overall Energy (eV)"
	for i in range(nval):
		if occ=="o" or occ=="O":
			ws[alphabet[i+2]+"1"]="HOMO-"+str(i)+" Energy (eV)"
		elif occ=="u" or occ=="U":
			ws[alphabet[i+2]+"1"]="LUMO+"+str(i)+" Energy (eV)"
		else:
			raise Exception("Use 'o' for occupied orbitals or 'u' for unoccupied")
	for i,center in enumerate(centers):
		column=nval+i+2
		ws[alphabet[column]+"1"]="Center "+str(i+1)+" ChElPG"

	mo_e, e_tots, chelpg = get_serial_energies(out_file,nval,occ)

	for i,frame in enumerate(e_tots):
			ws["A"+str(i+2)]=i
			ws["B"+str(i+2)]=frame
			for j in range(nval):
				if occ=="o" or occ=="O":
					ws[alphabet[j+2]+str(i+2)]=mo_e[i][j]
				elif occ=="u" or occ=="U":
					ws[alphabet[j+2]+str(i+2)]=mo_e[i][j]

	

	for i,frame in enumerate(chelpg):
		for j,atom in enumerate(frame):
			if j in centers:
				column=nval+j+2
				ws[alphabet[column]+str(i+2)]=float(chelpg[i][j])

	wb.save(excel)

def get_serial_energies(out_file,nval,occ):
	'''
	Takes a serial Q-Chem output file and returns energies of specified MOs, total energies, and ChElPG charges
	Inputs:
		out_file: str. the name of the qchem output file
		nval: int. number of orbitals to plot from band gap
		occ: str. either o for occupied or u for unoccupied. can enter guess and it will guess
	Outputs: 
		mo_e: np array of mo energies in eV. njobs x nval
		e_tots: np array of total energies in eV. njobs
		chelpg: np array of chelpg charges. njobs x natoms
	'''
	atoms,coords=get_geom_io(out_file)
	if occ.lower()=="guess":
		if "P" in atoms and ("In" in atoms or "Ga" in atoms):
			p_count=0
			m_count=0
			for i,atom in enumerate(atoms):
				if atom=="P":
					p_count=p_count+1
				if atom=="In" or atom=="Ga":
					m_count=m_count+1
			if p_count<m_count:
				occ="o"
			else:
				occ="u"
		elif "P" in atoms:
			occ="o"
		else:
			occ="u"


	with open(out_file,"r") as out:
		e_tots=[]
		all_mo_e=[]
		mo_block=False

		center_charges=[]
		charge_flag=False
		charge_count=0

		for i,line in enumerate(out):
			if line.find('Welcome to Q-Chem')!= -1:
				frame_mos=[]

				frame_charges=[]
				charge_count=0

			if line.find('Total energy in the final basis set')!= -1:
				tot_e=line.strip().split()[-1]
				e_tots.append(float(tot_e)*27.211396641308)
			if line.find('Thank you very much for using Q-Chem.  Have a nice day.') != -1:
				if occ=="o" or occ=="O":
					e_val_edge=frame_mos[-nval:]
				elif occ=="u" or occ=="U":
					e_val_edge=frame_mos[:nval]
				ev_edge=[float(x)*27.211396641308 for x in e_val_edge]
				all_mo_e.append(ev_edge)

				center_charges.append(frame_charges)
			if line.find("Ground-State ChElPG Net Atomic Charges")!=-1:
				charge_flag=True
			if charge_flag==True:
				charge_count=charge_count+1
			if charge_count>4 and line.find('----------------------------------------')!=-1:
				charge_flag=False
			if charge_flag==True and charge_count>4:
				frame_charges.append(line.strip().split())


			if occ=="o" or occ=="O":
				if line.find('-- Occupied --')!= -1 and mo_block==False:
					mo_block=True
				elif mo_block==True and line.find('-- Virtual --')!= -1:
					mo_block=False
				elif mo_block==True:
					frame_mos.extend(line.strip().split())
			elif occ=="u" or occ=="U":
				if line.find('-- Virtual --')!= -1 and mo_block==False:
					mo_block=True
				elif mo_block==True and line.find('--------------------------------------------------------------')!= -1:
					mo_block=False
				elif mo_block==True:
					frame_mos.extend(line.strip().split())

	mo_e=np.zeros((len(e_tots),nval))
	for i,frame in enumerate(e_tots):
		for j in range(nval):
			mo_e[i][j]=all_mo_e[i][-(j+1)]


	chelpg=np.zeros((len(center_charges),len(atoms)))
	for i,frame in enumerate(center_charges):
		for j,atom in enumerate(atoms):
			chelpg[i][j]=center_charges[i][j][2]

	e_tots=np.array(e_tots)
	

	return mo_e, e_tots, chelpg


if __name__ == "__main__":
	main()